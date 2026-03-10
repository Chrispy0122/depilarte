from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from backend.database import get_db
from . import models, services
from backend.modules.agenda import models as agenda_models
from backend.modules.pacientes import models as pacientes_models
from pydantic import BaseModel
from datetime import datetime, date, time, timedelta
from sqlalchemy import func, or_
from typing import Optional
from backend.modules.inventario import services as inventario_services
from backend.modules.inventario import models as inventario_models
import traceback # For debugging 500 errors

router = APIRouter(
    prefix="/api/cobranza",
    tags=["Cobranza"]
)

class PagoCreate(BaseModel):
    cita_id: int
    cliente_id: Optional[int] = None  # Required for flash appointments when cita_id=0
    monto_pagado: float
    metodo: str
    referencia: Optional[str] = None
    usar_wallet: bool = False
    abono_wallet: float = 0.0  # Extra amount to add to wallet
    proxima_cita: Optional[date] = None # New Field

@router.get("/tasa-bcv")
async def get_tasa_bcv():
    tasa = await services.obtener_tasa_bcv()
    return {"tasa": tasa}

@router.get("/pendientes")
def get_citas_por_cobrar(db: Session = Depends(get_db)):
    # 1. FECHA DE HOY (Venezuela Time: UTC-4)
    # Ajuste manual a Hora Venezuela (UTC - 4 horas)
    utc_now = datetime.utcnow()
    venezuela_now = utc_now - timedelta(hours=4)
    hoy_venezuela = venezuela_now.date()
    
    # Debug print (Optional, remove in prod)
    # print(f"Server Date: {date.today()} | Venezuela Date: {hoy_venezuela}")
    
    # 2. FILTRO AMPLIADO
    # - Fecha: HOY (Venezuela)
    # - Status: Confirmada o Asistio (Solo pacientes que ya están listos para pagar)
    # - Excluir: 'pagada', 'cancelada', 'pendiente', 'agendada'
    
    estados_activos = ['confirmada', 'asistio']
    
    citas = db.query(agenda_models.Cita).join(agenda_models.Cita.cliente).filter(
        func.date(agenda_models.Cita.fecha_hora_inicio) == hoy_venezuela,
        func.lower(agenda_models.Cita.estado).in_(estados_activos),
    ).all()
    
    resultados = []
    seen_clientes = set()
    
    for cita in citas:
        if cita.cliente_id in seen_clientes:
            continue
        seen_clientes.add(cita.cliente_id)

        # Determine wallet status
        tiene_wallet = False
        if cita.cliente and cita.cliente.saldo_wallet > 0:
            tiene_wallet = True

        # Calculate Retention suggestion
        proxima_cita_texto = "No definida"
        if cita.cliente and cita.cliente.frecuencia_visitas:
            fecha_sug = hoy_venezuela + timedelta(days=cita.cliente.frecuencia_visitas)
            # Format: "En 21 días (16 Feb)"
            # Note: strftime %b is locale dependent, sticking to simple format or english for now
            meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
            mes_str = meses[fecha_sug.month - 1]
            proxima_cita_texto = f"En {cita.cliente.frecuencia_visitas} días ({fecha_sug.day} {mes_str})"
            
        # NEW: Calculate monto from servicios (M:N relationship)
        monto_total = 0.0
        servicios_nombres = []
        if cita.servicios and len(cita.servicios) > 0:
            for servicio in cita.servicios:
                monto_total += servicio.precio_sesion
                servicios_nombres.append(servicio.nombre)
            servicio_str = ", ".join(servicios_nombres)
        else:
            servicio_str = "Sin servicio"
            
        resultados.append({
            "cita_id": cita.id,
            "cliente_id": cita.cliente_id,
            "paciente_nombre": cita.cliente.nombre_completo if cita.cliente else "Desconocido",
            "hora": cita.fecha_hora_inicio.strftime("%H:%M"),
            "servicio": servicio_str,  # NOW: Actual service names
            "monto_esperado": monto_total,  # NOW: Calculated from servicios
            "tiene_wallet": tiene_wallet,
            "proxima_cita_texto": proxima_cita_texto
        })
        
    return resultados

@router.get("/pacientes-del-dia")
def get_pacientes_del_dia(db: Session = Depends(get_db)):
    # ROLLBACK: Retornar TODOS los pacientes activos para permitir cobro libre.
    # Se elimina el filtro de fecha y status.
    
    pacientes = db.query(pacientes_models.Cliente).order_by(pacientes_models.Cliente.nombre_completo.asc()).limit(100).all()
    
    resultados = []
    for p in pacientes:
        resultados.append({
            "cita_id": 0, # No hay cita especifica vinculada en este modo "directorio"
            "cliente_id": p.id,
            "paciente_nombre": p.nombre_completo,
            "hora": "---", 
            "servicio": "General",
            "monto_esperado": 0.0
        })
        
    return resultados

@router.get("/")
def listar_pagos(db: Session = Depends(get_db)):
    # List recent payments with Client data
    pagos = db.query(models.Pago).\
        join(models.Pago.cita).\
        join(agenda_models.Cita.cliente).\
        order_by(models.Pago.id.desc()).limit(50).all()
        
    historial = []
    for pago in pagos:
        # User requested nested client object for frontend convenience
        cliente_obj = {
            "nombre": pago.cita.cliente.nombre_completo,
            "apellido": "", # DB stores full name in one field
            "id": pago.cita.cliente.id
        }
        
        historial.append({
            "id": pago.id,
            "fecha": pago.cita.fecha_hora_inicio.date(), 
            "cliente": cliente_obj, # Nested as requested
            "monto": pago.monto,
            "metodo": pago.metodo,
            "referencia": pago.referencia
        })
    return historial

@router.get("/exportar-excel")
def exportar_excel(db: Session = Depends(get_db)):
    """Genera un archivo Excel con el historial de cobros"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    # 1. Obtener datos (mismo query que listar_pagos pero sin limite o con limite mayor)
    pagos = db.query(models.Pago).\
        join(models.Pago.cita).\
        join(agenda_models.Cita.cliente).\
        order_by(models.Pago.id.desc()).limit(1000).all()
    
    # 2. Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Cobros"
    
    # Encabezados
    headers = ["ID Pago", "Fecha", "Cliente", "Cédula", "Monto ($)", "Método", "Referencia"]
    ws.append(headers)
    
    # Estilo Encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B7A58", end_color="2B7A58", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    # Datos
    for pago in pagos:
        ws.append([
            pago.id,
            pago.cita.fecha_hora_inicio.date(),
            pago.cita.cliente.nombre_completo,
            pago.cita.cliente.cedula,
            pago.monto,
            pago.metodo,
            pago.referencia or ""
        ])
        
    # Ajustar ancho columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    # 3. Guardar en memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # 4. Retornar respuesta
    filename = f"Cobros_Depilarte_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        excel_file, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


from .schemas import CobroCreate
from .models import Cobro, DetalleCobro, Pago

@router.post("/")
def crear_cobro(cobro_in: CobroCreate, db: Session = Depends(get_db)):
    """
    Registra una venta (Cobro) con múltiples items.
    Soporta precios override y tipos de venta.
    """
    
    # 1. Calcular total real (Precio de lista)
    try:
        grand_total = 0.0
        auto_wallet_topup = 0.0
        mensaje_extra = None

        for item in cobro_in.items:
            # INTERCEPT FULL PACKAGE
            if str(item.tipo_venta).lower() == 'paquete' and item.tipo_cobro == 'completo':
                # Split logic
                sesiones = item.sesiones_totales if item.sesiones_totales > 0 else 1
                precio_por_sesion = item.precio_aplicado / sesiones
                monto_abono = item.precio_aplicado - precio_por_sesion

                # Record the original cost so it's not lost when we overwrite
                # Adding a temporary property just to hold it during the loop
                setattr(item, '_costo_original_completo', item.precio_aplicado)

                # Overwrite applied price with only 1 session's worth
                item.precio_aplicado = precio_por_sesion
                auto_wallet_topup += monto_abono

                # Generate message if there's a split
                if monto_abono > 0:
                    mensaje_extra = f"Pago procesado. Se cobró ${precio_por_sesion:.2f} por la sesión de hoy y se abonaron ${monto_abono:.2f} al Wallet."

            # Calculate grand_total with the (potentially modified) applied price
            grand_total += item.precio_aplicado
            
        # --- LOGICA CORREGIDA: ABONO = WALLET TOP-UP ---
        # monto_total_venta = Precio de los servicios (lo que costó)
        # monto_wallet_usado = Lo que se pagó con saldo a favor
        # monto_abonado (Input) = Lo que se recargó EXTRA a la wallet
        
        wallet_used = cobro_in.monto_wallet_usado or 0.0
        
        # Include the auto-calculated split top-up
        wallet_topup = (cobro_in.monto_abonado or 0.0) + auto_wallet_topup
        
        cash_for_service = max(0.0, grand_total - wallet_used)
        total_cash_in = cash_for_service + wallet_topup
            
        # 1.5 Auto-assign Staff (Single Specialist/Receptionist Policy)
        from backend.modules.staff.models import Empleado
        
        spec_emp = db.query(Empleado).filter(or_(Empleado.rol == 'especialista', Empleado.rol == 'ambos'), Empleado.activo == 1).first()
        rec_emp = db.query(Empleado).filter(or_(Empleado.rol == 'recepcionista', Empleado.rol == 'ambos'), Empleado.activo == 1).first()
        
        spec_id = spec_emp.id if spec_emp else None
        rec_id = rec_emp.id if rec_emp else None

        # 2. Crear Header
        nuevo_cobro = Cobro(
            cliente_id=cobro_in.cliente_id,
            fecha=datetime.now(),
            metodo_pago=cobro_in.metodo_pago,
            referencia=cobro_in.referencia,
            # New Fields Corrected
            monto_total_venta=grand_total, # Full Price (Modified for 1 session if package)
            monto_abonado=wallet_topup,    # Wallet Top-Up (includes auto_wallet_topup)
            deuda=0.0,                     # No debt
            total=total_cash_in,           # Real Money In
            tasa_bcv=cobro_in.tasa_bcv     # Tasa BCV histórica del día
        )
        db.add(nuevo_cobro)
        db.flush() # Ensure nuevo_cobro gets its ID here without committing
        
        # 2.1 MANTENER HISTORIAL GENERAL (Cita + Pago)
        from backend.modules.agenda.models import Cita, EstadoCita
        from backend.modules.cobranza.models import Pago
        
        # Flash appointment for History tracking
        start_of_day = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)
        
        existing_cita = db.query(Cita).filter(
            Cita.cliente_id == cobro_in.cliente_id,
            Cita.fecha_hora_inicio >= start_of_day,
            Cita.fecha_hora_inicio < end_of_day
        ).first()
        
        if existing_cita:
            new_cita = existing_cita
            new_cita.estado = EstadoCita.PAGADA
        else:
            new_cita = Cita(
                cliente_id=cobro_in.cliente_id,
                fecha_hora_inicio=datetime.now(),
                fecha_hora_fin=datetime.now() + timedelta(minutes=30),
                estado=EstadoCita.CONFIRMADA
            )
            db.add(new_cita)
            db.flush() # Ensure new_cita gets an ID
        
        if cash_for_service > 0:
            pago_cash = Pago(
                cita_id=new_cita.id,
                monto=cash_for_service,
                metodo=cobro_in.metodo_pago,
                referencia=cobro_in.referencia
            )
            db.add(pago_cash)
            
        if wallet_used > 0:
            pago_wallet = Pago(
                cita_id=new_cita.id,
                monto=wallet_used,
                metodo="Monedero",
                referencia="Uso de Saldo a Favor"
            )
            db.add(pago_wallet)
        
        # 2.5 Update Wallet Logic
        # DEDUCT used amount
        if wallet_used > 0:
            cliente = db.query(pacientes_models.Cliente).filter(pacientes_models.Cliente.id == cobro_in.cliente_id).first()
            if cliente:
                cliente.saldo_wallet -= wallet_used
                
        # ADD Top-up amount
        if wallet_topup > 0:
            cliente = db.query(pacientes_models.Cliente).filter(pacientes_models.Cliente.id == cobro_in.cliente_id).first()
            if cliente:
                cliente.saldo_wallet += wallet_topup
        
        # 3. Crear Detalles
        from backend.modules.servicios.models import PaqueteSpa
        from backend.modules.pacientes.models import PaqueteCliente
        
        for item in cobro_in.items:
            # Get Service Name snapshot
            servicio = db.query(PaqueteSpa).filter(PaqueteSpa.id == item.servicio_id).first()
            nombre_servicio = servicio.nombre if servicio else "Servicio Eliminado"
            original_price = servicio.sesion if item.tipo_venta == 'sesion' else servicio.paquete_4_sesiones
            
            # --- COMMISSION LOGIC ---
            # Get fixed commission rates from Service
            comision_recepcionista = servicio.comision_recepcionista if servicio else 0.0
            comision_especialista = servicio.comision_especialista if servicio else 0.0
            
            # Determine actual amounts based on AUTOMATIC staff presence
            monto_com_recep = comision_recepcionista if rec_id else 0.0
            monto_com_esp = comision_especialista if spec_id else 0.0

            detalle = DetalleCobro(
                cobro_id=nuevo_cobro.id,
                servicio_id=item.servicio_id,
                servicio_nombre=nombre_servicio,
                tipo_venta=item.tipo_venta,
                precio_unitario=original_price or 0.0,
                precio_aplicado=item.precio_aplicado,
                cantidad=1,
                # Staff & Commissions (AUTO ASSIGNED)
                recepcionista_id=rec_id,
                especialista_id=spec_id,
                monto_comision_recepcionista=monto_com_recep,
                monto_comision_especialista=monto_com_esp
            )
            db.add(detalle)
            
            # --- CREATE PAQUETE_CLIENTE IF SOLD ---
            if str(item.tipo_venta).lower() == 'paquete':
                # Determine original package cost based on whether it was a 'completo' override
                precio_paquete_real = getattr(item, '_costo_original_completo', item.precio_aplicado)
                
                # Si es un fraccionado, el costo total del paquete real es item.precio_aplicado * item.sesiones_totales
                # Si es completo, el costo total es el precio_paquete_real directamente y ya entra pagado.
                costo_base = original_price if original_price else precio_paquete_real * item.sesiones_totales
                
                # Por seguridad, si viene completo, el costo total es lo que pagó
                if item.tipo_cobro == 'completo':
                    costo_real = precio_paquete_real
                else: 
                    # Es fraccionado, el costo_total pactado es [precio de esta cuota * total_sesiones] como proxy, o el original_price.  
                    # Usaremos el original_price (paquete_4_sesiones en BD) si existe, o proxy.
                    costo_real = original_price if original_price else (precio_paquete_real * item.sesiones_totales)
                
                nuevo_paquete_cliente = PaqueteCliente(
                    paciente_id=cobro_in.cliente_id,
                    nombre_paquete=nombre_servicio,
                    total_sesiones=item.sesiones_totales,
                    sesiones_usadas=1,                   # Se usa la primera sesión el día de la compra 
                    costo_total=costo_real,
                    monto_pagado=precio_paquete_real,    # Lo que pagó hoy legalmente por el paquete (fracción o completo)
                    activo=True if 1 < item.sesiones_totales else False
                )
                db.add(nuevo_paquete_cliente)

            # --- LOGIC PORT: INVENTORY CONSUMPTION ---
            # Consumir receta si existe (Inventory Logic)
            receta = inventario_services.obtener_receta_por_servicio(db, item.servicio_id)
            if receta:
                 inventario_services.consumir_receta(
                    db, 
                    receta.id, 
                    referencia=f"Cobro #{nuevo_cobro.id} - {nombre_servicio}"
                )

        # 4. Actualizar Próxima Cita (Legacy Logic)
        cliente = db.query(pacientes_models.Cliente).filter(pacientes_models.Cliente.id == cobro_in.cliente_id).first()
        if cliente:
            if cobro_in.fecha_proxima:
                # Parse YYYY-MM-DD
                try:
                    y, m, d = map(int, cobro_in.fecha_proxima.split('-'))
                    cliente.fecha_proxima_estimada = date(y, m, d)
                except:
                    pass # Ignore invalid date format
            else:
                # Auto
                frec = cliente.frecuencia_visitas if cliente.frecuencia_visitas else 21
                cliente.fecha_proxima_estimada = date.today() + timedelta(days=frec)

        db.commit()
        db.refresh(nuevo_cobro)
        
        return {
            "mensaje": "Cobro registrado correctamente",
            "mensaje_extra": mensaje_extra,
            "cobro_id": nuevo_cobro.id,
            "total": grand_total,
            "abonado": wallet_topup,
            "deuda": 0.0
        }
    except Exception as e:
        db.rollback()
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error procesando cobro: {str(e)}")

# --- LEGACY /pagar ENDPOINT KEPT FOR BACKWARD COMPAT IF NEEDED ---
@router.post("/pagar")
def procesar_pago(pago: PagoCreate, db: Session = Depends(get_db)):
    cita = None
    
    # 1. Handle "Flash Appointment" creation if no cita_id provided
    if not pago.cita_id or pago.cita_id == 0:
        if not pago.cliente_id:
             raise HTTPException(status_code=400, detail="Debe indicar cita_id o cliente_id")
        
        # Create ad-hoc appointment for this payment
        # This keeps the DB consistent (Payment -> Cita -> Client) without forcing the user to use the Calendar
        new_cita = agenda_models.Cita(
            cliente_id=pago.cliente_id,
            fecha_hora_inicio=datetime.now(),
            fecha_hora_fin=datetime.now() + timedelta(minutes=30), # Default duration
            estado=agenda_models.EstadoCita.CONFIRMADA # Auto-confirm
            # NOTE: No servicios assigned to flash appointments - monto comes from pago.monto_pagado
        )
        db.add(new_cita)
        db.commit()
        db.refresh(new_cita)
        cita = new_cita
        pago.cita_id = new_cita.id # Link payment to this new appointment
        
    else:
        # Existing Logic
        cita = db.query(agenda_models.Cita).filter(agenda_models.Cita.id == pago.cita_id).first()
        if not cita:
            raise HTTPException(status_code=404, detail="Cita no encontrada")

    # Buscar Cliente (from Cita, which ensures consistency)
    cliente = cita.cliente
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente asociado a la cita no encontrado")

    # Calcular monto total de la cita desde servicios (M:N)
    total_a_pagar = 0.0
    if cita.servicios and len(cita.servicios) > 0:
        for servicio in cita.servicios:
            total_a_pagar += servicio.precio_sesion
    
    # Fallback: if cita has no services, use the amount being paid
    if total_a_pagar <= 0: 
        total_a_pagar = pago.monto_pagado

    saldo_usado_wallet = 0.0
    
    # 1. Lógica Wallet (Descuento)
    if pago.usar_wallet and cliente.saldo_wallet > 0:
        # Logic remains same: deduct available from total
        if cliente.saldo_wallet >= total_a_pagar:
            saldo_usado_wallet = total_a_pagar
        else:
            saldo_usado_wallet = cliente.saldo_wallet
            
    # El monto que el usuario PAGA REALMENTE (Cash/Zelle) + Wallet usado
    total_cubierto = saldo_usado_wallet + pago.monto_pagado
    
    # 2. Lógica Overpayment (A abonar a Wallet) - SOLO para sobrepago no intencional
    if total_cubierto > total_a_pagar:
        excedente = total_cubierto - total_a_pagar
        cliente.saldo_wallet += excedente
    
    # 3. Lógica Abono Intencional a Wallet
    if pago.abono_wallet > 0:
        cliente.saldo_wallet += pago.abono_wallet
    
    # 4. Descontar del Wallet lo usado
    if saldo_usado_wallet > 0:
        cliente.saldo_wallet -= saldo_usado_wallet

    # 4. ACTUALIZAR FECHA PRÓXIMA CITA
    # If provided manually, use it
    if pago.proxima_cita:
        cliente.fecha_proxima_estimada = pago.proxima_cita
    else:
        # AUTO-CALCULATE from frecuencia_visitas
        # Default: 21 days if not set
        frecuencia = cliente.frecuencia_visitas if cliente.frecuencia_visitas else 21
        cliente.fecha_proxima_estimada = date.today() + timedelta(days=frecuencia)

    # Registrar el Pago Real
    nuevo_pago = models.Pago(
        cita_id=cita.id,
        monto=pago.monto_pagado, 
        metodo=pago.metodo,
        referencia=pago.referencia
    )
    db.add(nuevo_pago)
    
    # Registrar consumo de Wallet si hubo
    if saldo_usado_wallet > 0:
        pago_wallet = models.Pago(
            cita_id=cita.id,
            monto=saldo_usado_wallet,
            metodo="WALLET",
            referencia=f"Wallet debit: {saldo_usado_wallet}"
        )
        db.add(pago_wallet)

    # Actualizar estado a PAGADA para que no vuelva a aparecer en pendientes
    if total_cubierto >= total_a_pagar:
        cita.estado = agenda_models.EstadoCita.PAGADA
        
        # --- CONSUMO AUTOMÁTICO DE INVENTARIO (BOM) ---
        if cita.servicios:
            for servicio in cita.servicios:
                # Buscar receta activa para el servicio
                receta = inventario_services.obtener_receta_por_servicio(db, servicio.id)
                if receta:
                    inventario_services.consumir_receta(
                        db, 
                        receta.id, 
                        referencia=f"Cita #{cita.id} - {servicio.nombre}"
                    )

    db.commit()
    return {
        "mensaje": "Pago procesado exitosamente",
        "total_pagado_real": pago.monto_pagado,
        "wallet_usado": saldo_usado_wallet,
        "cita_id_generada": cita.id
    }

@router.get("/hoy")
def obtener_cobros_hoy(
    fecha: str = None, # Optional YYYY-MM-DD
    db: Session = Depends(get_db)
):
    """
    Retorna los cobros realizados en una fecha específica.
    Si no se envía fecha, retorna los de HOY.
    """
    try:
        if fecha:
            try:
                hoy = datetime.strptime(fecha, "%Y-%m-%d").date()
            except ValueError:
                 raise HTTPException(status_code=400, detail="Formato de fecha inválido. Usar YYYY-MM-DD")
        else:
            hoy = datetime.now().date()
            
        start_of_day = datetime.combine(hoy, time.min)
        end_of_day = datetime.combine(hoy, time.max)
        
        # print(f"DEBUG API CAJA: Querying {start_of_day} to {end_of_day}")
        
        cobros = db.query(Cobro).filter(
            Cobro.fecha >= start_of_day,
            Cobro.fecha <= end_of_day
        ).order_by(Cobro.fecha.desc()).all()
        
        # Calculate daily totals
        # Total Real Money In = Sum(c.total) which now equals (ServiceCash + TopUp)
        total_dia = sum((c.total or 0.0) for c in cobros)
        total_comisiones = 0.0
        
        detalle_cobros = []
        for c in cobros:
            # Calculate commissions for this cobro
            comision_cobro = 0.0
            msg_servicios = []
            for det in c.detalles:
                comision_cobro += (det.monto_comision_recepcionista or 0.0) + (det.monto_comision_especialista or 0.0)
                msg_servicios.append(det.servicio_nombre or "Servicio")
                
            total_comisiones += comision_cobro
            
            # Use new meanings
            # monto_total_venta = Service Price (List Price)
            # monto_abonado = Wallet Top-Up
            # total = Total Cash In (Service Cash + Top Up)
            
            cash_in = c.total or 0.0
            wallet_topup = c.monto_abonado or 0.0
            cash_service = max(0.0, cash_in - wallet_topup)
            
            detalle_cobros.append({
                "id": c.id,
                "hora": c.fecha.strftime("%H:%M"),
                "cliente": c.cliente.nombre_completo if c.cliente else "Desconocido",
                "monto_total": cash_service, # VISUAL FIX: Show Cash Paid for Service
                "monto_abonado": wallet_topup,   # Wallet Top-Up
                "deuda": 0.0, 
                "monto": cash_in, # Cash In (Total)
                "metodo": c.metodo_pago,
                "servicios": ", ".join(msg_servicios),
                "referencia": c.referencia or "-",
                "comisiones_generadas": comision_cobro,
                "tasa_bcv": c.tasa_bcv or None  # Tasa BCV histórica
            })
            
        return {
            "fecha": hoy.isoformat(),
            "total_cobrado": total_dia,
            "total_comisiones": total_comisiones,
            "cobros": detalle_cobros
        }
    except Exception as e:
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)} | Trace: {traceback.format_exc()}")

@router.get("/nomina/historico")
def obtener_nomina_historico(
    start_date: str, # YYYY-MM-DD
    end_date: str,   # YYYY-MM-DD
    db: Session = Depends(get_db)
):
    """
    Reporte de Nómina REESCRITO (Strict Mode).
    Estrategia: Dos consultas separadas para evitar conflictos de agregación.
    Fusion: Python Dictionary Merge.
    """
    try:
        f_inicio = datetime.strptime(start_date, "%Y-%m-%d")
        f_fin = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Usar YYYY-MM-DD")

    from backend.modules.staff.models import Empleado
    from backend.modules.servicios.models import PaqueteSpa
    import re
    
    # --- HELPER FUNCTIONS ---
    def parse_zonas(zonas_str):
        if not zonas_str: return 0
        try:
            # "4 a 5" -> [4, 5] -> max 5
            nums = [int(n) for n in re.findall(r'\d+', str(zonas_str))]
            return max(nums) if nums else 0
        except:
            return 0
            
    nomina = {} 
    
    def get_or_create_emp(emp_id):
        if emp_id not in nomina:
            # Fetch Employee Metadata ONCE
            emp = db.query(Empleado).filter(Empleado.id == emp_id).first()
            if emp:
                nomina[emp_id] = {
                    "empleado_id": emp.id,
                    "nombre": emp.nombre_completo,
                    "rol": emp.rol, # Roles can be 'ambos', so this is just their default role
                    "total_pagar": 0.0,
                    "total_zonas": 0,
                    "total_limpiezas": 0,
                    "total_masajes": 0,
                    "detalle_items": [] # Optional debug
                }
        return nomina.get(emp_id)

    # --- PASO 1: RECEPCIONISTAS ---
    # Traer todos los items donde hay recepcionista asignada
    detalles_rec = db.query(DetalleCobro).join(Cobro).filter(
        Cobro.fecha >= f_inicio,
        Cobro.fecha <= f_fin,
        DetalleCobro.recepcionista_id != None
    ).all()
    
    for d in detalles_rec:
        emp_data = get_or_create_emp(d.recepcionista_id)
        if not emp_data: continue
        
        # 1. Sumar Dinero (EXCLUSIVAMENTE Recep)
        emp_data["total_pagar"] += (d.monto_comision_recepcionista or 0.0)
        
        # 2. Sumar Volumen
        svc = d.servicio # Lazy load ok
        if svc:
            cat = str(svc.categoria).lower().strip()
            
            if cat == "depilacion":
                z = parse_zonas(svc.num_zonas)
                emp_data["total_zonas"] += z
                
            elif cat == "facial":
                emp_data["total_limpiezas"] += 1
                
            elif cat == "corporal":
                emp_data["total_masajes"] += 1

    # --- PASO 2: ESPECIALISTAS ---
    # Traer todos los items donde hay especialista asignada
    detalles_spec = db.query(DetalleCobro).join(Cobro).filter(
        Cobro.fecha >= f_inicio,
        Cobro.fecha <= f_fin,
        DetalleCobro.especialista_id != None
    ).all()
    
    for d in detalles_spec:
        emp_data = get_or_create_emp(d.especialista_id)
        if not emp_data: continue
        
        # 1. Sumar Dinero (EXCLUSIVAMENTE Spec)
        emp_data["total_pagar"] += (d.monto_comision_especialista or 0.0)
        
        # 2. Sumar Volumen (Mismo servicio, pero cuenta para el especialista tambien)
        svc = d.servicio
        if svc:
            cat = str(svc.categoria).lower().strip()
            
            if cat == "depilacion":
                z = parse_zonas(svc.num_zonas)
                emp_data["total_zonas"] += z
                
            elif cat == "facial":
                emp_data["total_limpiezas"] += 1
                
            elif cat == "corporal":
                emp_data["total_masajes"] += 1

    # --- RETURN ---
    return list(nomina.values())

@router.get("/exportar")
def exportar_caja_excel(
    fecha: str = None, 
    db: Session = Depends(get_db)
):
    """
    Genera y descarga un Excel con el reporte de Caja de una fecha.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from fastapi.responses import StreamingResponse

        # 1. Determine Date
        if fecha:
            try:
                target_date = datetime.strptime(fecha, "%Y-%m-%d").date()
            except:
                raise HTTPException(status_code=400, detail="Fecha inválida")
        else:
            target_date = datetime.now().date()
            
        # 2. Query Data
        start = datetime.combine(target_date, time.min)
        end = datetime.combine(target_date, time.max)
        
        cobros = db.query(Cobro).filter(
            Cobro.fecha >= start,
            Cobro.fecha <= end
        ).order_by(Cobro.fecha.asc()).all()
        
        # 3. Build Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Caja {target_date}"

        # ── Helpers de estilo ──────────────────────────────
        from openpyxl.utils import get_column_letter

        def make_border(style='thin'):
            s = Side(style=style)
            return Border(left=s, right=s, top=s, bottom=s)

        def cell_style(ws, row, col, value=None, bold=False, font_color="000000",
                       bg=None, align="left", wrap=False, border=True, num_fmt=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=bold, color=font_color, name="Calibri", size=10)
            if bg:
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
            if border:
                c.border = make_border()
            if num_fmt:
                c.number_format = num_fmt
            return c

        # ── Colores de la paleta ───────────────────────────
        CLR_GREEN_DARK  = "2B7A58"   # verde Depilarte
        CLR_GREEN_LIGHT = "E8F5E9"   # fondo verde suave
        CLR_BLUE_DARK   = "1565C0"   # azul para Bs
        CLR_BLUE_LIGHT  = "E3F2FD"   # fondo azul suave
        CLR_GRAY_DARK   = "455A64"   # gris oscuro
        CLR_GRAY_LIGHT  = "ECEFF1"   # fondo gris suave
        CLR_WHITE       = "FFFFFF"
        CLR_TITLE_BG    = "1B4F3A"   # verde muy oscuro para título
        CLR_ROW_ALT     = "F5F5F5"   # alternancia de filas

        TOTAL_COLS = 12

        # ── FILA 1: Título principal ───────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "DEPILARTE  ·  Reporte de Caja Diaria"
        title_cell.font = Font(bold=True, color=CLR_WHITE, name="Calibri", size=14)
        title_cell.fill = PatternFill(start_color=CLR_TITLE_BG, end_color=CLR_TITLE_BG, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # ── FILA 2: Subtítulo con fecha ────────────────────
        meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        fecha_str = f"{target_date.day} de {meses[target_date.month-1]} de {target_date.year}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
        sub_cell = ws.cell(row=2, column=1)
        sub_cell.value = f"Fecha: {fecha_str}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        sub_cell.font = Font(italic=True, color=CLR_WHITE, name="Calibri", size=10)
        sub_cell.fill = PatternFill(start_color=CLR_GREEN_DARK, end_color=CLR_GREEN_DARK, fill_type="solid")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # ── FILA 3: Vacía de separación ────────────────────
        ws.row_dimensions[3].height = 6

        # ── FILA 4: Encabezados de columnas por sección ───
        # Sección INFO (cols 1-5): verde oscuro
        # Sección USD  (cols 6-7): verde medio
        # Sección Bs   (cols 8-9): azul
        # Sección BCV  (col 10):   gris
        # Sección Staff(cols 11-12): gris oscuro

        header_defs = [
            # (col, texto, bg, alineación)
            (1,  "Hora",                 CLR_GREEN_DARK, "center"),
            (2,  "Cliente",              CLR_GREEN_DARK, "left"),
            (3,  "Servicios",            CLR_GREEN_DARK, "left"),
            (4,  "Método de Pago",       CLR_GREEN_DARK, "center"),
            (5,  "Referencia",           CLR_GREEN_DARK, "center"),
            (6,  "Total Servicio ($)",   "217A4E",       "right"),
            (7,  "Abono Wallet ($)",     "217A4E",       "right"),
            (8,  "Total Servicio (Bs)",  CLR_BLUE_DARK,  "right"),
            (9,  "Abono Wallet (Bs)",    CLR_BLUE_DARK,  "right"),
            (10, "Tasa BCV",             CLR_GRAY_DARK,  "center"),
            (11, "Recepcionista",        CLR_GRAY_DARK,  "center"),
            (12, "Especialista",         CLR_GRAY_DARK,  "center"),
        ]

        ws.row_dimensions[4].height = 22
        for col, text, bg, align in header_defs:
            c = ws.cell(row=4, column=col, value=text)
            c.font = Font(bold=True, color=CLR_WHITE, name="Calibri", size=10)
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            c.border = make_border()

        # ── FILAS DE DATOS ─────────────────────────────────
        total_usd = 0.0
        total_bs  = 0.0

        for idx, c in enumerate(cobros):
            data_row = 5 + idx

            # Calcular valores
            cash_in      = c.total or 0.0
            wallet_topup = c.monto_abonado or 0.0
            cash_service = max(0.0, cash_in - wallet_topup)
            tasa         = c.tasa_bcv

            cash_service_bs  = round(cash_service  * tasa, 2) if tasa else None
            wallet_topup_bs  = round(wallet_topup  * tasa, 2) if tasa else None

            total_usd += cash_service
            if cash_service_bs is not None:
                total_bs += cash_service_bs

            # Nombres servicios / personal
            svcs   = []
            receps = set()
            specs  = set()
            for d in c.detalles:
                svcs.append(d.servicio_nombre or "?")
                if d.recepcionista: receps.add(d.recepcionista.nombre_completo)
                if d.especialista:  specs.add(d.especialista.nombre_completo)

            # Color de fila alternado
            row_bg = CLR_ROW_ALT if idx % 2 == 0 else CLR_WHITE

            # Columnas INFO
            for col, val, align, wrap in [
                (1, c.fecha.strftime("%H:%M"),                       "center", False),
                (2, c.cliente.nombre_completo if c.cliente else "?", "left",   False),
                (3, ", ".join(svcs),                                 "left",   True ),
                (4, c.metodo_pago,                                   "center", False),
                (5, c.referencia or "-",                             "center", False),
            ]:
                cell_style(ws, data_row, col, val, bg=row_bg, align=align, wrap=wrap)

            # Columnas USD
            for col, val in [(6, cash_service), (7, wallet_topup)]:
                cell_style(ws, data_row, col, val,
                           bg=CLR_GREEN_LIGHT, align="right",
                           num_fmt='"$"#,##0.00')

            # Columnas Bs
            for col, val in [(8, cash_service_bs), (9, wallet_topup_bs)]:
                cell_style(ws, data_row, col, val,
                           bg=CLR_BLUE_LIGHT, align="right",
                           num_fmt='"Bs "#,##0.00' if val is not None else None)

            # Tasa BCV
            cell_style(ws, data_row, 10, tasa,
                       bg=CLR_GRAY_LIGHT, align="center",
                       num_fmt="#,##0.00" if tasa else None)

            # Personal
            for col, val in [(11, ", ".join(receps) or "-"), (12, ", ".join(specs) or "-")]:
                cell_style(ws, data_row, col, val, bg=CLR_GRAY_LIGHT, align="center")

            ws.row_dimensions[data_row].height = 18

        # ── FILA TOTALES ───────────────────────────────────
        total_row = 5 + len(cobros)
        ws.row_dimensions[total_row].height = 22

        # Merge etiqueta
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
        label = ws.cell(row=total_row, column=1, value="TOTAL DEL DÍA")
        label.font = Font(bold=True, color=CLR_WHITE, name="Calibri", size=11)
        label.fill = PatternFill(start_color=CLR_TITLE_BG, end_color=CLR_TITLE_BG, fill_type="solid")
        label.alignment = Alignment(horizontal="right", vertical="center")
        label.border = make_border()

        # Total USD
        cell_style(ws, total_row, 6, total_usd, bold=True,
                   font_color=CLR_WHITE, bg=CLR_GREEN_DARK, align="right",
                   num_fmt='"$"#,##0.00')
        # Vacío Wallet $
        cell_style(ws, total_row, 7, None, bg=CLR_GREEN_DARK, border=True)

        # Total Bs
        cell_style(ws, total_row, 8, total_bs if total_bs > 0 else None,
                   bold=True, font_color=CLR_WHITE, bg=CLR_BLUE_DARK, align="right",
                   num_fmt='"Bs "#,##0.00' if total_bs > 0 else None)
        cell_style(ws, total_row, 9, None, bg=CLR_BLUE_DARK, border=True)

        # Celdas vacías finales fila totales
        for col in [10, 11, 12]:
            cell_style(ws, total_row, col, None, bg=CLR_GRAY_LIGHT, border=True)

        # ── ANCHOS DE COLUMNA ──────────────────────────────
        col_widths = {1:8, 2:26, 3:38, 4:15, 5:16,
                      6:18, 7:18, 8:20, 9:20, 10:11, 11:20, 12:20}
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Congelar encabezados (fila 4)
        ws.freeze_panes = "A5"

        
        # Save to Buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Caja_Depilarte_{target_date}.xlsx"
        
        return StreamingResponse(
            buffer, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"Export Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# PAQUETES / CUPONERA  (lógica financiera concentrada en Cobranza)
# ═══════════════════════════════════════════════════════════════════════════════
from .schemas import PaqueteClienteCreate, PaqueteClienteOut, AbonarSesionIn
from typing import List as TypingList

@router.get("/paciente/{paciente_id}/paquetes", response_model=TypingList[PaqueteClienteOut])
def listar_paquetes_paciente(paciente_id: int, db: Session = Depends(get_db)):
    """Retorna todos los paquetes activos del paciente (usado por Cobranza y por Pacientes en lectura)."""
    cliente = db.query(pacientes_models.Cliente).filter(pacientes_models.Cliente.id == paciente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")
    return db.query(pacientes_models.PaqueteCliente).filter(
        pacientes_models.PaqueteCliente.paciente_id == paciente_id,
        pacientes_models.PaqueteCliente.activo == True
    ).all()


@router.post("/paciente/{paciente_id}/paquetes", response_model=PaqueteClienteOut)
def vender_paquete(paciente_id: int, data: PaqueteClienteCreate, db: Session = Depends(get_db)):
    """Vende (crea) un nuevo paquete de sesiones al paciente."""
    cliente = db.query(pacientes_models.Cliente).filter(pacientes_models.Cliente.id == paciente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")
    nuevo = pacientes_models.PaqueteCliente(
        paciente_id=paciente_id,
        nombre_paquete=data.nombre_paquete,
        total_sesiones=data.total_sesiones,
        sesiones_usadas=0,
        costo_total=data.costo_total,
        monto_pagado=0.0,
        activo=True,
    )
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo


@router.post("/paquete/{paquete_id}/abonar", response_model=PaqueteClienteOut)
def abonar_sesion_paquete(paquete_id: int, data: AbonarSesionIn, db: Session = Depends(get_db)):
    """
    Registra el pago de UNA sesión del paquete:
    - Suma 1 a sesiones_usadas.
    - Suma el monto a monto_pagado.
    - Marca el paquete como inactivo si se agotaron las sesiones.
    El cobro también debe registrarse en la caja llamando al endpoint POST /api/cobranza/.
    """
    paquete = db.query(pacientes_models.PaqueteCliente).filter(
        pacientes_models.PaqueteCliente.id == paquete_id
    ).first()
    if not paquete:
        raise HTTPException(status_code=404, detail="Paquete no encontrado")
    if not paquete.activo:
        raise HTTPException(status_code=409, detail="El paquete ya está completado o inactivo")
    if paquete.sesiones_usadas >= paquete.total_sesiones:
        raise HTTPException(status_code=409, detail="Ya se usaron todas las sesiones de este paquete")

    paquete.sesiones_usadas += 1
    paquete.monto_pagado = round(paquete.monto_pagado + data.monto, 2)

    # Auto-cerrar si se completaron todas las sesiones
    if paquete.sesiones_usadas >= paquete.total_sesiones:
        paquete.activo = False

    db.commit()
    db.refresh(paquete)
    return paquete
